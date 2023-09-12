import openpyxl


# read data
def read_data_from_spreadsheet(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    data = list()

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    workbook.close()
    return data


# write data
def write_data_to_spreadsheet(filename, sheet_name, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    # clear existing data - optional
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    # write new data
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(filename)
    workbook.close()


input_filename = 'DependentFilesForLab_9b/Book1.xlsx'
output_filename = 'DependentFilesForLab_9b/Book1.xlsx'
input_sheet_name = 'Sheet1'
output_sheet_name = 'Sheet1'


data_to_read = read_data_from_spreadsheet(input_filename, input_sheet_name)
print("Data read from the spreadsheet: ")
for row in data_to_read:
    print(row)


modified_data = [[cell + 1 for cell in row] for row in data_to_read]
write_data_to_spreadsheet(output_filename, output_sheet_name, modified_data)
print(f"Modified data written to {output_filename} in the {output_sheet_name} sheet")
